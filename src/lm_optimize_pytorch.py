import os
import sys
import numpy as np
import torch
import csv

# 添加上级目录到Python路径，以便导入tools模块
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)
# 添加当前目录到Python路径，以便导入同目录下的模块
sys.path.append(current_dir)

from tools.data_loader import (
    load_joint_angles, get_initial_params, get_error_weights,
    get_fixed_indices, get_parameter_groups, get_optimizable_indices,
    get_laser_tool_matrix,
    get_alternate_optimization_config, get_damping_config, 
    get_convergence_config, get_constraints_config,
    get_output_config, get_max_theta_change_radians
)
from jacobian_torch import (
    compute_error_vector_jacobian, 
    forward_kinematics_T, 
    quaternion_to_rotation_matrix
)

#! 将旋转矩阵转换为四元数
def _rotation_matrix_to_quaternion(R_matrix):
    """ 
    旋转矩阵转换为四元数公式：
        R = [rx, ry, rz]
        q = [qx, qy, qz, qw]
        qx = (R[2,1] - R[1,2]) / S
        qy = (R[0,2] - R[2,0]) / S
        qz = (R[1,0] - R[0,1]) / S
        qw = (R[0,0] + R[1,1] + R[2,2] + 1) / 4
    """
    if not torch.is_tensor(R_matrix):
        R_matrix = torch.as_tensor(R_matrix, dtype=torch.float64)
    q = torch.zeros(4, dtype=R_matrix.dtype, device=R_matrix.device)
    trace = R_matrix[0,0] + R_matrix[1,1] + R_matrix[2,2]   
    if trace > 1e-8: 
        S = torch.sqrt(trace + 1.0) * 2.0 
        q[3] = 0.25 * S  
        q[0] = (R_matrix[2,1] - R_matrix[1,2]) / S 
        q[1] = (R_matrix[0,2] - R_matrix[2,0]) / S 
        q[2] = (R_matrix[1,0] - R_matrix[0,1]) / S 
    elif (R_matrix[0,0] > R_matrix[1,1]) and (R_matrix[0,0] > R_matrix[2,2]):
        S = torch.sqrt(1.0 + R_matrix[0,0] - R_matrix[1,1] - R_matrix[2,2] + 1e-12) * 2.0 
        q[0] = 0.25 * S # qx
        q[1] = (R_matrix[0,1] + R_matrix[1,0]) / S 
        q[2] = (R_matrix[0,2] + R_matrix[2,0]) / S 
    elif R_matrix[1,1] > R_matrix[2,2]:
        S = torch.sqrt(1.0 + R_matrix[1,1] - R_matrix[0,0] - R_matrix[2,2] + 1e-12) * 2.0 
        q[3] = (R_matrix[0,2] - R_matrix[2,0]) / S 
        q[0] = (R_matrix[0,1] + R_matrix[1,0]) / S 
        q[1] = 0.25 * S 
        q[2] = (R_matrix[1,2] + R_matrix[2,1]) / S 
    else:
        S = torch.sqrt(1.0 + R_matrix[2,2] - R_matrix[0,0] - R_matrix[1,1] + 1e-12) * 2.0 
        q[3] = (R_matrix[1,0] - R_matrix[0,1]) / S 
        q[0] = (R_matrix[0,2] + R_matrix[2,0]) / S 
        q[1] = (R_matrix[1,2] + R_matrix[2,1]) / S 
        q[2] = 0.25 * S 

    #* 归一化四元数
    norm_q = torch.linalg.norm(q)
    if norm_q > 1e-9: 
        q = q / norm_q
    else: 
        q = torch.tensor([0.0, 0.0, 0.0, 1.0], dtype=R_matrix.dtype, device=R_matrix.device)
    return q 

#! 将四元数转换为欧拉角 (ZYX顺序: yaw, pitch, roll)
def _quaternion_to_euler_angles(q):
    """ 
    四元数转换为欧拉角公式：
        q = [qx, qy, qz, qw]
        roll = atan2(sinr_cosp, cosr_cosp)
        pitch = asin(sinp)
        yaw = atan2(siny_cosp, cosy_cosp)
    """
    qx, qy, qz, qw = q[0], q[1], q[2], q[3]

 
    sinr_cosp = 2 * (qw * qx + qy * qz)
    cosr_cosp = 1 - 2 * (qx * qx + qy * qy)
    roll = torch.atan2(sinr_cosp, cosr_cosp)

    sinp = 2 * (qw * qy - qz * qx)
    if torch.abs(sinp) >= 1:
        pitch = torch.copysign(torch.pi / 2, sinp) 
    else:
        pitch = torch.asin(sinp)

   
    siny_cosp = 2 * (qw * qz + qx * qy)
    cosy_cosp = 1 - 2 * (qy * qy + qz * qz)
    yaw = torch.atan2(siny_cosp, cosy_cosp)

    return torch.stack([yaw, pitch, roll])

#! 计算单组数据的误差向量（加权重）
def compute_error_vector(params, joint_angles, laser_matrix, weights=None):
    # 如果没有提供权重，从配置获取
    if weights is None:
        weights = get_error_weights()
        
    #* 获取关节角度和参数
    q_t = torch.as_tensor(joint_angles, dtype=torch.float64)
    params_t = torch.as_tensor(params, dtype=torch.float64) 

    #* 提取参数
    params_for_fk = params_t[0:31] 
    t_laser_base_pos = params_t[31:34]
    t_laser_base_quat = params_t[34:38] 

  
    T_pred_robot_base, _ = forward_kinematics_T(q_t, params_for_fk) 

    #* 构建T_laser_base变换矩阵（基座在激光坐标系下的位姿）
    R_laser_base = quaternion_to_rotation_matrix(t_laser_base_quat) 
    T_laser_base_matrix = torch.eye(4, dtype=torch.float64)
    T_laser_base_matrix[0:3, 0:3] = R_laser_base
    T_laser_base_matrix[0:3, 3] = t_laser_base_pos
    
    #* 将机器人预测位姿转换到激光跟踪仪坐标系
    T_pred_in_laser_frame = torch.matmul(T_laser_base_matrix, T_pred_robot_base)
    
    #* 计算激光跟踪仪-基座变换矩阵
    pred_pos = T_pred_in_laser_frame[0:3, 3]
    pred_R = T_pred_in_laser_frame[0:3, 0:3]

    #* 提取测量位置和旋转矩阵
    T_laser_t = torch.as_tensor(laser_matrix, dtype=torch.float64)
    meas_pos = T_laser_t[0:3, 3]
    meas_R = T_laser_t[0:3, 0:3]

    #* 位置误差
    pos_error = pred_pos - meas_pos

    #* 将旋转矩阵转换为四元数 [qx, qy, qz, qw]
    q_pred = _rotation_matrix_to_quaternion(pred_R) 
    q_meas = _rotation_matrix_to_quaternion(meas_R) 

    #* 计算四元数共轭
    q_meas_conj_x = -q_meas[0]
    q_meas_conj_y = -q_meas[1]
    q_meas_conj_z = -q_meas[2]
    q_meas_conj_w =  q_meas[3]
    
    # 四元数乘法: q_pred * q_meas_conj
    # q_pred = [x1, y1, z1, w1], q_meas_conj = [x2, y2, z2, w2]
    # qw = w1*w2 - x1*x2 - y1*y2 - z1*z2
    # qx = w1*x2 + x1*w2 + y1*z2 - z1*y2
    # qy = w1*y2 - x1*z2 + y1*w2 + z1*x2
    # qz = w1*z2 + x1*y2 - y1*x2 + z1*w2
    #* 计算预测和实际之间的误差旋转
    q_err_w = q_pred[3] * q_meas_conj_w - q_pred[0] * q_meas_conj_x - q_pred[1] * q_meas_conj_y - q_pred[2] * q_meas_conj_z
    q_err_x = q_pred[3] * q_meas_conj_x + q_pred[0] * q_meas_conj_w + q_pred[1] * q_meas_conj_z - q_pred[2] * q_meas_conj_y
    q_err_y = q_pred[3] * q_meas_conj_y - q_pred[0] * q_meas_conj_z + q_pred[1] * q_meas_conj_w + q_pred[2] * q_meas_conj_x
    q_err_z = q_pred[3] * q_meas_conj_z + q_pred[0] * q_meas_conj_y - q_pred[1] * q_meas_conj_x + q_pred[2] * q_meas_conj_w

    #* 将误差四元数转换为欧拉角误差 [yaw, pitch, roll]
    error_quaternion = torch.stack([q_err_x, q_err_y, q_err_z, q_err_w])
    # 归一化误差四元数
    norm_error_q = torch.linalg.norm(error_quaternion)
    if norm_error_q > 1e-9:
        error_quaternion_normalized = error_quaternion / norm_error_q
    else:
        # 如果模长接近于零，说明旋转很小，或者q_pred和q_meas非常接近，欧拉角误差也接近于零
        error_quaternion_normalized = torch.tensor([0.0, 0.0, 0.0, 1.0], dtype=params.dtype, device=params.device)

    orient_error = _quaternion_to_euler_angles(error_quaternion_normalized) # 使用欧拉角作为方向误差
    
    combined_error = torch.cat((pos_error, orient_error))
    return combined_error * torch.as_tensor(weights, dtype=torch.float64)



#! 计算所有样本的均方根误差
def compute_total_error_avg(params, joint_angles, laser_matrices, weights=None):
    # 如果没有提供权重，从配置获取
    if weights is None:
        weights = get_error_weights()
        
    total_error_sum_sq = 0.0 
    n_samples = len(joint_angles)
    if n_samples == 0:
        return torch.tensor(0.0, dtype=torch.float64) 

    for i in range(n_samples):
        error_vec = compute_error_vector(params, joint_angles[i], laser_matrices[i], weights) 
        total_error_sum_sq += torch.sum(error_vec**2)
    
    #* 返回均方根误差 (RMSE误差) 公式：RMSE = sqrt(sum(error_vec^2) / n_samples)
    mean_squared_error = total_error_sum_sq / n_samples
    return torch.sqrt(mean_squared_error)

#! 生成优化前后详细对比分析
def generate_detailed_comparison(initial_params, optimized_params):
    """
    生成优化前后的详细对比分析数据
    
    参数:
    initial_params: 初始参数
    optimized_params: 优化后的参数
    
    返回:
    pandas.DataFrame: 包含详细对比数据的DataFrame
    """
    from tools.data_loader import load_joint_angles, get_laser_tool_matrix
    import pandas as pd
    
    # 加载测试数据
    joint_angles = load_joint_angles()
    laser_matrices = get_laser_tool_matrix()
    n_samples = len(joint_angles)
    
    print(f"   📊 分析 {n_samples} 组测试数据的优化效果...")
    
    # 计算权重
    weights = get_error_weights()
    
    comparison_data = []
    
    for i in range(n_samples):
        joint_angle = joint_angles[i]
        laser_matrix = laser_matrices[i]
        
        # 计算优化前的误差
        error_before = compute_error_vector(initial_params, joint_angle, laser_matrix, weights)
        pos_error_before = error_before[:3].detach().numpy()
        orient_error_before = error_before[3:].detach().numpy()
        total_error_before = torch.norm(error_before).item()
        
        # 计算优化前的分解误差（位置误差和姿态误差）
        pos_error_magnitude_before = np.sqrt(np.sum(pos_error_before**2))
        orient_error_magnitude_before = np.sqrt(np.sum(orient_error_before**2))
        
        # 计算优化后的误差
        error_after = compute_error_vector(optimized_params, joint_angle, laser_matrix, weights)
        pos_error_after = error_after[:3].detach().numpy()
        orient_error_after = error_after[3:].detach().numpy()
        total_error_after = torch.norm(error_after).item()
        
        # 计算优化后的分解误差（位置误差和姿态误差）
        pos_error_magnitude_after = np.sqrt(np.sum(pos_error_after**2))
        orient_error_magnitude_after = np.sqrt(np.sum(orient_error_after**2))
        
        # 计算改进率
        improvement_rate = ((total_error_before - total_error_after) / total_error_before) * 100 if total_error_before > 0 else 0
        pos_improvement_rate = ((pos_error_magnitude_before - pos_error_magnitude_after) / pos_error_magnitude_before) * 100 if pos_error_magnitude_before > 0 else 0
        orient_improvement_rate = ((orient_error_magnitude_before - orient_error_magnitude_after) / orient_error_magnitude_before) * 100 if orient_error_magnitude_before > 0 else 0
        
        # 添加到对比数据 - 按照优化前后成对的顺序排列
        comparison_data.append({
            '数据组': f'第{i+1}组',
            '优化前X误差(mm)': f'{pos_error_before[0]:.6f}',
            '优化后X误差(mm)': f'{pos_error_after[0]:.6f}',
            '优化前Y误差(mm)': f'{pos_error_before[1]:.6f}',
            '优化后Y误差(mm)': f'{pos_error_after[1]:.6f}',
            '优化前Z误差(mm)': f'{pos_error_before[2]:.6f}',
            '优化后Z误差(mm)': f'{pos_error_after[2]:.6f}',
            '优化前Rx误差(度)': f'{orient_error_before[0]:.6f}',
            '优化后Rx误差(度)': f'{orient_error_after[0]:.6f}',
            '优化前Ry误差(度)': f'{orient_error_before[1]:.6f}',
            '优化后Ry误差(度)': f'{orient_error_after[1]:.6f}',
            '优化前Rz误差(度)': f'{orient_error_before[2]:.6f}',
            '优化后Rz误差(度)': f'{orient_error_after[2]:.6f}',
            '优化前位置误差(mm)': f'{pos_error_magnitude_before:.6f}',
            '优化后位置误差(mm)': f'{pos_error_magnitude_after:.6f}',
            '优化前姿态误差(度)': f'{orient_error_magnitude_before:.6f}',
            '优化后姿态误差(度)': f'{orient_error_magnitude_after:.6f}',
            '优化前总误差(L2范数)': f'{total_error_before:.6f}',
            '优化后总误差(L2范数)': f'{total_error_after:.6f}',
            '位置改进率(%)': f'{pos_improvement_rate:.2f}%',
            '姿态改进率(%)': f'{orient_improvement_rate:.2f}%',
            '总误差改进率(%)': f'{improvement_rate:.2f}%'
        })
    
    # 计算统计信息
    total_errors_before = [float(row['优化前总误差(L2范数)']) for row in comparison_data]
    total_errors_after = [float(row['优化后总误差(L2范数)']) for row in comparison_data]
    pos_errors_before = [float(row['优化前位置误差(mm)']) for row in comparison_data]
    pos_errors_after = [float(row['优化后位置误差(mm)']) for row in comparison_data]
    orient_errors_before = [float(row['优化前姿态误差(度)']) for row in comparison_data]
    orient_errors_after = [float(row['优化后姿态误差(度)']) for row in comparison_data]
    
    avg_error_before = np.mean(total_errors_before)
    avg_error_after = np.mean(total_errors_after)
    avg_pos_error_before = np.mean(pos_errors_before)
    avg_pos_error_after = np.mean(pos_errors_after)
    avg_orient_error_before = np.mean(orient_errors_before)
    avg_orient_error_after = np.mean(orient_errors_after)
    
    overall_improvement = ((avg_error_before - avg_error_after) / avg_error_before) * 100 if avg_error_before > 0 else 0
    pos_overall_improvement = ((avg_pos_error_before - avg_pos_error_after) / avg_pos_error_before) * 100 if avg_pos_error_before > 0 else 0
    orient_overall_improvement = ((avg_orient_error_before - avg_orient_error_after) / avg_orient_error_before) * 100 if avg_orient_error_before > 0 else 0
    
    # 添加统计行 - 匹配新的列顺序
    comparison_data.append({
        '数据组': '平均值',
        '优化前X误差(mm)': '',
        '优化后X误差(mm)': '',
        '优化前Y误差(mm)': '',
        '优化后Y误差(mm)': '',
        '优化前Z误差(mm)': '',
        '优化后Z误差(mm)': '',
        '优化前Rx误差(度)': '',
        '优化后Rx误差(度)': '',
        '优化前Ry误差(度)': '',
        '优化后Ry误差(度)': '',
        '优化前Rz误差(度)': '',
        '优化后Rz误差(度)': '',
        '优化前位置误差(mm)': f'{avg_pos_error_before:.6f}',
        '优化后位置误差(mm)': f'{avg_pos_error_after:.6f}',
        '优化前姿态误差(度)': f'{avg_orient_error_before:.6f}',
        '优化后姿态误差(度)': f'{avg_orient_error_after:.6f}',
        '优化前总误差(L2范数)': f'{avg_error_before:.6f}',
        '优化后总误差(L2范数)': f'{avg_error_after:.6f}',
        '位置改进率(%)': f'{pos_overall_improvement:.2f}%',
        '姿态改进率(%)': f'{orient_overall_improvement:.2f}%',
        '总误差改进率(%)': f'{overall_improvement:.2f}%'
    })
    
    print(f"   ✅ 总体平均误差: {avg_error_before:.6f} → {avg_error_after:.6f}")
    print(f"   📍 位置误差: {avg_pos_error_before:.6f} → {avg_pos_error_after:.6f} (改进率: {pos_overall_improvement:.2f}%)")
    print(f"   🔄 姿态误差: {avg_orient_error_before:.6f} → {avg_orient_error_after:.6f} (改进率: {orient_overall_improvement:.2f}%)")
    print(f"   📈 总体改进率: {overall_improvement:.2f}%")
    
    return pd.DataFrame(comparison_data)

#! 保存优化后的DH参数和TCP参数
def save_optimization_results(params, initial_params=None, filepath_prefix=None):
    """
    保存优化结果，路径可以从配置文件读取
    
    参数:
    params: 优化后的参数
    initial_params: 初始参数（用于对比分析）
    filepath_prefix: 文件路径前缀（None时从配置读取）
    """
    # 从配置文件读取默认路径
    if filepath_prefix is None:
        output_config = get_output_config()
        filepath_prefix = output_config.get('results_prefix', 'results/optimized')
    
    dirpath = os.path.dirname(filepath_prefix)
    if dirpath and not os.path.exists(dirpath):
        os.makedirs(dirpath)
    dh_params = params[0:24]
    tcp_params = params[24:31]
    t_laser_base_params = params[31:38]
    dh_filepath = f"{filepath_prefix}_dh_parameters.csv"
    dh_matrix = np.array(dh_params).reshape(6, 4)
    header_dh = "alpha,a,d,theta_offset"
    row_labels_dh = [f"Joint_{i+1}" for i in range(6)]
    with open(dh_filepath, 'w') as f:
        f.write(f",{header_dh}\n")  
        for i, row in enumerate(dh_matrix):
            f.write(f"{row_labels_dh[i]},{row[0]:.6f},{row[1]:.6f},{row[2]:.6f},{row[3]:.6f}\n")
    print(f"优化后的DH参数已保存到: {dh_filepath}")
    
    # 保存TCP参数
    tcp_filepath = f"{filepath_prefix}_tcp_parameters.csv"
    header_tcp = "parameter,value"
    tcp_param_names = ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]
    with open(tcp_filepath, 'w') as f:
        f.write(f"{header_tcp}\n")
        for name, value in zip(tcp_param_names, tcp_params):
            f.write(f"{name},{value:.6f}\n")
    print(f"优化后的TCP参数已保存到: {tcp_filepath}")
    
    # 保存激光跟踪仪-基座变换参数
    t_laser_base_filepath = f"{filepath_prefix}_t_laser_base_parameters.csv"
    header_t_laser_base = "parameter,value"
    t_laser_base_param_names = ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]
    with open(t_laser_base_filepath, 'w') as f:
        f.write(f"{header_t_laser_base}\n")
        for name, value in zip(t_laser_base_param_names, t_laser_base_params):
            f.write(f"{name},{value:.6f}\n")
    print(f"优化后的激光跟踪仪-基座变换参数已保存到: {t_laser_base_filepath}")
    
    # 生成详细对比分析数据
    detailed_comparison_df = None
    if initial_params is not None:
        print("\n🔍 生成优化前后详细对比分析...")
        detailed_comparison_df = generate_detailed_comparison(initial_params, params)
    
    # 自动生成Excel汇总文件
    try:
        import pandas as pd
        from pathlib import Path
        
        # 创建Excel汇总文件
        results_dir = Path(dirpath)
        excel_filepath = results_dir / "优化结果汇总.xlsx"
        
        # 读取三个CSV文件的数据
        dh_df = pd.read_csv(dh_filepath)
        tcp_df = pd.read_csv(tcp_filepath)
        base_df = pd.read_csv(t_laser_base_filepath)
        
        # 创建一个综合的数据框，将三个结果合并到一个工作表
        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
            # 创建工作簿
            workbook = writer.book
            worksheet = workbook.create_sheet(title="优化结果汇总")
            
            # 设置标题样式
            from openpyxl.styles import Font, Alignment, PatternFill
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=12)
            center_alignment = Alignment(horizontal='center', vertical='center')
            header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            
            # 写入DH参数部分
            row = 1
            worksheet.cell(row=row, column=1, value="🔧 机器人DH参数标定结果").font = title_font
            worksheet.cell(row=row, column=1).alignment = center_alignment
            worksheet.merge_cells(f'A{row}:E{row}')
            row += 2
            
            # DH参数表头
            dh_headers = ['关节编号', 'α (度)', 'a (mm)', 'd (mm)', 'θ偏移 (度)']
            for col, header in enumerate(dh_headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            row += 1
            
            # DH参数数据
            for i, (_, dh_row) in enumerate(dh_df.iterrows()):
                worksheet.cell(row=row, column=1, value=dh_row.iloc[0])  # 关节名称
                for col in range(1, 5):
                    worksheet.cell(row=row, column=col+1, value=f"{dh_row.iloc[col]:.6f}")
                row += 1
            
            # 空行分隔
            row += 2
            
            # 写入TCP位姿部分
            worksheet.cell(row=row, column=1, value="🎯 工具中心点(TCP)位姿").font = title_font
            worksheet.cell(row=row, column=1).alignment = center_alignment
            worksheet.merge_cells(f'A{row}:B{row}')
            row += 2
            
            # TCP位姿表头和数据
            tcp_headers = ['位姿参数', '优化值']
            for col, header in enumerate(tcp_headers, 1):
                cell = worksheet.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            row += 1
            
            # TCP参数名称美化映射
            tcp_param_names = {
                'tx': 'X轴位移 (mm)',
                'ty': 'Y轴位移 (mm)', 
                'tz': 'Z轴位移 (mm)',
                'qx': '四元数 qx',
                'qy': '四元数 qy',
                'qz': '四元数 qz',
                'qw': '四元数 qw'
            }
            
            for _, tcp_row in tcp_df.iterrows():
                param_name = tcp_param_names.get(tcp_row['parameter'], tcp_row['parameter'])
                worksheet.cell(row=row, column=1, value=param_name)
                worksheet.cell(row=row, column=2, value=f"{tcp_row['value']:.6f}")
                row += 1
            
            # 写入基座位姿部分（与TCP位姿并排显示）
            row_start_base = row - len(tcp_df) - 1  # 回到TCP标题行
            worksheet.cell(row=row_start_base, column=4, value="📍 激光跟踪仪基座位姿").font = title_font
            worksheet.cell(row=row_start_base, column=4).alignment = center_alignment
            worksheet.merge_cells(f'D{row_start_base}:E{row_start_base}')
            row_base = row_start_base + 2
            
            # 基座位姿表头
            base_headers = ['位姿参数', '优化值']
            for col, header in enumerate(base_headers, 4):
                cell = worksheet.cell(row=row_base, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            row_base += 1
            
            # 基座参数名称美化映射
            base_param_names = {
                'tx': 'X轴位移 (mm)',
                'ty': 'Y轴位移 (mm)',
                'tz': 'Z轴位移 (mm)', 
                'qx': '四元数 qx',
                'qy': '四元数 qy',
                'qz': '四元数 qz',
                'qw': '四元数 qw'
            }
            
            # 基座位姿数据
            for _, base_row in base_df.iterrows():
                param_name = base_param_names.get(base_row['parameter'], base_row['parameter'])
                worksheet.cell(row=row_base, column=4, value=param_name)
                worksheet.cell(row=row_base, column=5, value=f"{base_row['value']:.6f}")
                row_base += 1
            
            # 调整列宽
            worksheet.column_dimensions['A'].width = 18
            worksheet.column_dimensions['B'].width = 15
            worksheet.column_dimensions['C'].width = 15
            worksheet.column_dimensions['D'].width = 20
            worksheet.column_dimensions['E'].width = 15
            
            # 在主工作表底部添加说明
            last_row = worksheet.max_row + 2
            worksheet.cell(row=last_row, column=1, value="📋 说明").font = title_font
            worksheet.cell(row=last_row, column=1).alignment = center_alignment
            
            explanation_lines = [
                "• DH参数：机器人正向运动学标定参数",
                "• TCP位姿：工具中心点相对法兰坐标系的变换",
                "• 基座位姿：激光跟踪仪与机器人基座间的变换关系",
                "• 详细误差分析请查看'优化前后详细对比'工作表"
            ]
            
            for i, line in enumerate(explanation_lines):
                worksheet.cell(row=last_row + 1 + i, column=1, value=line).font = Font(size=10)
                worksheet.merge_cells(f'A{last_row + 1 + i}:E{last_row + 1 + i}')
            
            # 添加边框样式
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 为所有有数据的单元格添加边框
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
            
            # 添加详细对比分析工作表
            if detailed_comparison_df is not None:
                # 创建对比分析工作表
                comparison_sheet = workbook.create_sheet(title="优化前后详细对比")
                
                                # 写入标题
                comparison_sheet.cell(row=1, column=1, value="📊 优化前后逐组数据对比分析").font = title_font
                comparison_sheet.cell(row=1, column=1).alignment = center_alignment
                comparison_sheet.merge_cells('A1:V1')  # 扩展合并范围以适应新列
                
                # 添加计算说明
                explanation_text = ("💡 误差计算方式:\n"
                                  "   位置误差 = √(X²+Y²+Z²) (mm)\n"
                                  "   姿态误差 = √(Rx²+Ry²+Rz²) (度)\n" 
                                  "   总误差(L2范数) = √(X²+Y²+Z²+(Rx×0.01)²+(Ry×0.01)²+(Rz×0.01)²)\n"
                                  "   权重设置: 位置权重=1.0, 姿态权重=0.01")
                comparison_sheet.cell(row=2, column=1, value=explanation_text).font = Font(size=10, italic=True)
                comparison_sheet.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                comparison_sheet.merge_cells('A2:V2')  # 扩展合并范围以适应新列
                comparison_sheet.row_dimensions[2].height = 40
                
                # 写入表头
                headers = detailed_comparison_df.columns.tolist()
                for col, header in enumerate(headers, 1):
                    cell = comparison_sheet.cell(row=4, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # 写入数据
                for row_idx, (_, row_data) in enumerate(detailed_comparison_df.iterrows(), 5):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = comparison_sheet.cell(row=row_idx, column=col_idx, value=value)
                        # 为平均值行添加特殊样式
                        if '平均值' in str(value):
                            cell.font = header_font
                            cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
                
                # 调整列宽
                for col in range(1, len(headers) + 1):
                    if col == 1:  # 数据组列
                        comparison_sheet.column_dimensions[chr(64 + col)].width = 12
                    elif '改进率' in headers[col-1]:  # 改进率列
                        comparison_sheet.column_dimensions[chr(64 + col)].width = 14
                    elif '位置误差' in headers[col-1] or '姿态误差' in headers[col-1]:  # 位置/姿态误差列
                        comparison_sheet.column_dimensions[chr(64 + col)].width = 18
                    else:  # 其他误差列
                        comparison_sheet.column_dimensions[chr(64 + col)].width = 16
                
                # 为所有有数据的单元格添加边框
                for row in comparison_sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell.border = thin_border
            
            # 删除默认的Sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        print(f"✅ 优化结果汇总Excel文件已生成: {excel_filepath}")
        print("📊 Excel文件包含完整的优化分析:")
        print("   🔧 机器人DH参数标定结果 - 6个关节的完整DH参数")
        print("   🎯 工具中心点(TCP)位姿 - 位移+四元数表示") 
        print("   📍 激光跟踪仪基座位姿 - 基座坐标系变换参数")
        if detailed_comparison_df is not None:
            print("   📊 优化前后详细对比 - 逐组数据误差分析和改进率")
        print("   ✨ 专业格式化：图标标识、单位标注、边框美化")
        print("   📋 多工作表展示，全面分析优化效果")
        
    except ImportError:
        print("⚠️  未安装pandas库，无法生成Excel汇总文件")
        print("   请运行: pip install pandas openpyxl")
    except Exception as e:
        print(f"⚠️  生成Excel汇总文件时出错: {e}")

#! 使用增广系统SVD求解LM问题
def solve_lm_augmented_svd(J_opt, error_vector, lambda_val, damping_type=None, svd_threshold=None, verbose=None):
    """
    使用增广系统SVD求解Levenberg-Marquardt问题
    
    支持两种阻尼形式：
    1. Marquardt阻尼：H = J^T*J + λ * diag(J^T*J) （适合多尺度参数）
    2. Levenberg阻尼：H = J^T*J + λ * I （适合同尺度参数）
    
    参数:
    J_opt: 雅可比矩阵（只包含可优化参数的列）
    error_vector: 误差向量  
    lambda_val: 阻尼因子
    damping_type: "marquardt" 或 "levenberg"，阻尼类型（None时使用默认值）
    svd_threshold: 奇异值阈值，小于此值的奇异值被置零（None时使用默认值）
    verbose: 是否打印SVD诊断信息（None时使用默认值）
    
    返回:
    delta: 参数更新量
    svd_info: SVD诊断信息字典
    """
    # 使用固定默认值
    if damping_type is None:
        damping_config = get_damping_config()
        damping_type = damping_config.get('damping_type', 'marquardt')
    if svd_threshold is None:
        svd_threshold = 1e-12  # 固定默认值
    if verbose is None:
        verbose = True  # 固定默认值
    
    n_residuals, n_params = J_opt.shape
    lambda_tensor = torch.tensor(lambda_val, dtype=J_opt.dtype, device=J_opt.device)
    
    if damping_type == "marquardt":
        # Marquardt阻尼：H = J^T*J + λ * diag(J^T*J)
        # 增广系统等价形式：[J; √(λ * diag(J^T*J))]
        JTJ = torch.matmul(J_opt.transpose(0, 1), J_opt)
        diag_JTJ = torch.diag(JTJ)
        # 避免对角元素为零的情况（保护数值稳定性）
        diag_JTJ = torch.where(diag_JTJ > 1e-12, diag_JTJ, torch.tensor(1e-12, dtype=J_opt.dtype, device=J_opt.device))
        sqrt_diag = torch.sqrt(lambda_tensor * diag_JTJ)
        regularization_matrix = torch.diag(sqrt_diag)
    else:  # levenberg阻尼
        # Levenberg阻尼：使用单位矩阵
        sqrt_lambda = torch.sqrt(lambda_tensor)
        regularization_matrix = sqrt_lambda * torch.eye(n_params, dtype=J_opt.dtype, device=J_opt.device)
    
    # 构建增广雅可比矩阵：[J; D^(1/2)]
    J_aug = torch.vstack([J_opt, regularization_matrix])
    
    # 构建增广误差向量：[-e; 0] （注意负号，因为我们求解的是 min ||J*δ + e||²）
    zero_vector = torch.zeros(n_params, dtype=error_vector.dtype, device=error_vector.device)
    e_aug = torch.cat([-error_vector, zero_vector])
    
    # SVD分解
    U, S, Vt = torch.linalg.svd(J_aug, full_matrices=False)
    
    # 处理奇异值：设置阈值避免数值不稳定
    S_inv = torch.where(S > svd_threshold, 1.0/S, 0.0)
    effective_rank = torch.sum(S > svd_threshold).item()
    condition_number = S[0].item() / S[effective_rank-1].item() if effective_rank > 0 else float('inf')
    
    # 求解delta = V * S^(-1) * U^T * e_aug
    delta = Vt.T @ torch.diag(S_inv) @ U.T @ e_aug
    
    # 构建诊断信息
    svd_info = {
        'max_singular_value': S[0].item(),
        'min_singular_value': S[-1].item(),
        'effective_rank': effective_rank,
        'total_rank': len(S),
        'condition_number': condition_number,
        'lambda_val': lambda_val
    }
    
    if verbose:
        print(f"    SVD诊断: 有效秩={effective_rank}/{len(S)}, 条件数={condition_number:.2e}, "
              f"λ={lambda_val:.2e}")
    
    return delta, svd_info

#! 保存delta值到CSV文件
def save_delta_to_csv(delta, iteration, opt_indices, csv_file, lambda_val=None, error_val=None, alt_iteration=None, opt_step=None):
    """
    将delta值保存到CSV文件
    
    参数:
    delta: 参数更新量
    iteration: 当前迭代次数
    opt_indices: 可优化参数的索引
    csv_file: CSV文件路径
    lambda_val: 当前阻尼因子值（可选）
    error_val: 当前误差值（可选）
    alt_iteration: 交替优化循环次数（可选）
    opt_step: 优化步骤（1或2，可选）
    """
    try:
        # 创建目录（如果不存在）
        csv_dir = os.path.dirname(csv_file)
        if csv_dir and not os.path.exists(csv_dir):
            os.makedirs(csv_dir)
            
        # 创建完整的delta数组（38个参数）
        full_delta = np.zeros(38)
        for i, idx in enumerate(opt_indices):
            full_delta[idx] = delta[i]
        
        # 写入CSV文件
        with open(csv_file, 'a', newline='') as f:
            writer = csv.writer(f)
            row = []
            # 只添加delta值，不添加迭代信息、lambda和误差值
            row.extend(full_delta)
            writer.writerow(row)
    except Exception as e:
        print(f"保存delta值到CSV文件时出错: {e}")

#! LM优化
def optimize_dh_parameters(initial_params, max_iterations=None, lambda_init=None, tol=None, opt_indices=None, max_theta_delta_rad=None, csv_file=None, alt_iteration=None, opt_step=None):
    """
    LM优化函数，主要用于交替优化的子步骤
    
    参数:
    initial_params: 初始参数值
    max_iterations: 最大迭代次数（None时使用默认值50）
    lambda_init: 初始阻尼因子（None时从配置读取）
    tol: 收敛阈值（None时从配置读取）
    opt_indices: 可优化参数索引（None时优化所有非固定参数）
    max_theta_delta_rad: theta参数最大变化量（None时从配置读取）
    csv_file: CSV输出文件（None时从配置读取）
    alt_iteration: 交替优化轮次
    opt_step: 优化步骤
    """
    # 从配置文件读取默认参数
    damping_config = get_damping_config()
    convergence_config = get_convergence_config()
    constraints_config = get_constraints_config()
    output_config = get_output_config()
    
    # 使用传入参数或配置/默认值
    if max_iterations is None:
        max_iterations = 50  # 固定默认值，适合子优化步骤
    if lambda_init is None:
        lambda_init = damping_config.get('lambda_init_default', 0.01)
    if tol is None:
        tol = convergence_config.get('parameter_tol', 1e-10)
    if max_theta_delta_rad is None:
        max_theta_delta_rad = get_max_theta_change_radians()
    if csv_file is None and output_config.get('save_delta_values', True):
        csv_file = output_config.get('delta_csv_file', 'results/delta_values.csv')
    
    # 获取其他配置参数
    max_inner_iterations = convergence_config.get('max_inner_iterations', 10)
    rho_threshold = convergence_config.get('rho_threshold', 0.0)
    lambda_max = damping_config.get('lambda_max', 1e8)
    lambda_min = damping_config.get('lambda_min', 1e-7)
    enable_quat_norm = constraints_config.get('enable_quaternion_normalization', True)
    
    params = torch.tensor(initial_params, dtype=torch.float64, requires_grad=False)
    #* 初始化阻尼因子和加速因子
    lambda_val = lambda_init
    v_increase = 2  # 用于失败时增大lambda的加速因子
    consecutive_failures = 0  # 连续失败次数计数器
    max_consecutive_failures = 3  # 连续失败次数阈值
    
    #* 读取关节角度和激光数据
    joint_angles = load_joint_angles()
    laser_matrices = get_laser_tool_matrix()
    n_samples = len(joint_angles)
    if n_samples == 0:
        print("错误: 无法加载关节角度或激光数据，样本数量为0。")
        return initial_params
    
    #* 记录初始均方根误差
    current_avg_error = compute_total_error_avg(params, joint_angles, laser_matrices) 
    print(f"初始均方根误差：{current_avg_error.item():.6f}")
    avg_initial_error = current_avg_error.item() 
    
    #* 处理可优化参数索引
    if opt_indices is None:
        opt_indices = list(range(len(initial_params)))
    opt_indices = np.array(opt_indices)
    
    #* LM迭代
    for iteration in range(max_iterations):
        all_errors = []
        all_jacobians = []

        #* 计算所有样本的误差向量和雅可比矩阵
        for i in range(n_samples):
            error_vec = compute_error_vector(params, joint_angles[i], laser_matrices[i])
            jacobian = compute_error_vector_jacobian(params.numpy(), joint_angles[i], laser_matrices[i])
            all_errors.append(error_vec)
            all_jacobians.append(jacobian)
        error_vector = torch.cat(all_errors)

        #* 将所有雅可比矩阵堆叠成一个矩阵
        J = torch.vstack(all_jacobians)

        #* 使用增广SVD方法求解LM问题：更稳定的数值方法
        J_opt = J[:, opt_indices]
        update_success = False
        inner_iterations = 0
        while not update_success and inner_iterations < max_inner_iterations:
            inner_iterations += 1
            
            #! 使用增广SVD求解delta（从配置读取阻尼类型）
            try:
                delta, svd_info = solve_lm_augmented_svd(J_opt, error_vector, lambda_val)
                
                # 检查SVD求解质量
                if svd_info['effective_rank'] < len(opt_indices) * 0.8:
                    print(f"警告: 有效秩({svd_info['effective_rank']})较低，可能存在参数冗余")
                
                # 使用固定的条件数阈值
                condition_threshold = 1e12  # 固定默认值
                if svd_info['condition_number'] > condition_threshold:
                    print(f"警告: 条件数很大({svd_info['condition_number']:.2e})，增大阻尼因子")
                    lambda_val *= 5
                    if lambda_val > lambda_max:
                        print(f"阻尼因子超过阈值 {lambda_max}，提前结束优化")
                        return params.numpy()
                    continue

            except Exception as e:
                print(f"SVD求解错误: {e}，增大阻尼因子 λ = {lambda_val} -> {lambda_val * 10}")
                lambda_val *= 10
                if lambda_val > lambda_max:
                    print(f"阻尼因子超过阈值 {lambda_max}，提前结束优化")
                    return params.numpy()
                continue

            # 应用参数更新限制（针对theta角）
            if max_theta_delta_rad is not None and max_theta_delta_rad > 0:
                theta_param_indices_in_full_params = [3, 7, 11, 15, 19, 23]  # DH中theta_offset的索引
                for i, param_idx_in_full_params_np_val in enumerate(opt_indices):
                    param_idx_in_full_params = int(param_idx_in_full_params_np_val)  # 将numpy类型转换为int
                    if param_idx_in_full_params in theta_param_indices_in_full_params:
                        current_delta_val = delta[i]
                        # 将特定参数的更新量delta[i]限制在 [-max_theta_delta_rad, +max_theta_delta_rad] 范围内
                        delta[i] = torch.clamp(current_delta_val, -max_theta_delta_rad, max_theta_delta_rad)

            #* 尝试更新
            params_new = params.clone()
            params_new[opt_indices] += delta
            
            #* 计算新误差
            new_avg_error = compute_total_error_avg(params_new, joint_angles, laser_matrices)
            
            #* 使用rho策略进行lambda更新
            # 计算实际误差减少量（基于误差平方和）
            old_error_squared = current_avg_error.item() ** 2 * n_samples  # 转换回总误差平方和
            new_error_squared = new_avg_error.item() ** 2 * n_samples
            actual_reduction = old_error_squared - new_error_squared
            
            # 计算预测误差减少量: delta^T * (lambda * delta + J^T * error_vector)
            JTe = torch.matmul(J_opt.transpose(0, 1), error_vector)  # 重新计算J^T * e
            grad = JTe  # J^T * error_vector 
            predicted_reduction = torch.dot(delta, lambda_val * delta - grad).item()  # 注意这里用-grad因为JTe=-grad
            
            # 计算rho（增益比）
            if abs(predicted_reduction) > 1e-12:  # 避免除零
                rho = actual_reduction / predicted_reduction
            else:
                rho = -1.0  # 强制拒绝更新
            
            print(f"  rho = {rho:.4f}, 实际减少: {actual_reduction:.6f}, 预测减少: {predicted_reduction:.6f}")
            
            if rho > rho_threshold:  # 接受更新
                params = params_new
                current_avg_error = new_avg_error
                consecutive_failures = 0  # 重置连续失败计数器
                
                # 使用Nielsen策略更新lambda: lambda *= max(1/3, 1-(2*rho-1)^3)
                tmp = 2 * rho - 1
                factor = max(1.0/3.0, 1 - tmp * tmp * tmp)
                lambda_val = lambda_val * factor
                lambda_val = max(lambda_val, lambda_min)  # 使用配置的下界
                
                update_success = True
                print(f"  ✅ 接受更新, lambda: {lambda_val/factor:.4e} -> {lambda_val:.4e} (×{factor:.3f})")
            else:  # 拒绝更新
                consecutive_failures += 1  # 增加连续失败计数器
                
                # 增大lambda，使用加速惩罚
                lambda_val = lambda_val * v_increase
                old_v = v_increase
                v_increase = v_increase * 2  # 加速因子加倍
                update_success = False
                print(f"  ❌ 拒绝更新, lambda: {lambda_val/old_v:.4e} -> {lambda_val:.4e} (×{old_v}), 连续失败次数: {consecutive_failures}")
                
                # 检查连续失败次数是否达到阈值
                if consecutive_failures >= max_consecutive_failures:
                    print(f"连续失败 {consecutive_failures} 次，达到阈值 {max_consecutive_failures}，提前结束优化")
                    return params.numpy()
            
            # 如果接受了更新，重置加速因子
            if update_success:
                v_increase = 2

                #* TCP四元数归一化 
                if enable_quat_norm and any(idx in opt_indices for idx in range(27, 31)):
                    q_tcp = params[27:31] 
                    norm_q_tcp = torch.linalg.norm(q_tcp)
                    if norm_q_tcp > 1e-9: 
                        params[27:31] = q_tcp / norm_q_tcp
                    else:
                        params[27:31] = torch.tensor([0.0, 0.0, 0.0, 1.0], dtype=params.dtype, device=params.device)
                        print("警告: TCP四元数模长接近于零，已重置为[0,0,0,1]")
                
                #* 激光跟踪仪-基座四元数归一化 
                if enable_quat_norm and any(idx in opt_indices for idx in range(34, 38)):
                    q_laser_base = params[34:38]
                    norm_q_laser_base = torch.linalg.norm(q_laser_base)
                    if norm_q_laser_base > 1e-9:
                        params[34:38] = q_laser_base / norm_q_laser_base
                    else:
                        params[34:38] = torch.tensor([0.0, 0.0, 0.0, 1.0], dtype=params.dtype, device=params.device)
                        print("警告: 激光跟踪仪-基座四元数模长接近于零，已重置为[0,0,0,1]")

                # 保存delta值到CSV文件
                if csv_file is not None:
                    save_delta_to_csv(delta.numpy(), iteration+1, opt_indices, csv_file, 
                                     lambda_val=lambda_val, error_val=current_avg_error.item(), 
                                     alt_iteration=alt_iteration, opt_step=opt_step)

                print(f"迭代 {iteration+1}: 均方根误差 = {current_avg_error.item():.8f}, λ = {lambda_val:.4e}, \nΔθ (参数改变量) = {delta.numpy()}")            

            #* 如果阻尼因子超过一定阈值，提前结束优化
            if lambda_val > lambda_max * 0.1:  # 使用配置的90%作为警告阈值
                print(f"阻尼因子超过阈值 {lambda_max * 0.1:.0e}，提前结束优化")
                return params.numpy()
        if not update_success:
            print("内部迭代未收敛，继续主循环")
  
        if update_success and torch.norm(delta) < tol:
            print(f"参数变化小于阈值 {tol}，在第 {iteration+1} 次迭代后收敛")
            break

    #* 最终均方根误差
    final_avg_error = current_avg_error.item() 
    improvement = (avg_initial_error - final_avg_error) / avg_initial_error * 100 if avg_initial_error > 1e-9 else 0 
    print(f"优化完成，初始均方根误差: {avg_initial_error:.6f}, 最终均方根误差: {final_avg_error:.6f}, 改进率: {improvement:.2f}%")
    return params.numpy()

#! 交替优化函数
def alternate_optimize_parameters(initial_params, max_alt_iterations=None, convergence_tol=None, 
                                 max_sub_iterations_group1=None, max_sub_iterations_group2=None,
                                 lambda_init_group1=None, lambda_init_group2=None, max_theta_delta_rad_for_sub_opt=None):
    """
    交替优化函数，参数可以从配置文件读取默认值
    
    参数:
    initial_params: 初始参数值
    max_alt_iterations: 最大交替迭代次数（None时从配置读取）
    convergence_tol: 收敛阈值（None时从配置读取）
    max_sub_iterations_group1: 第一组参数最大迭代次数（None时从配置读取）
    max_sub_iterations_group2: 第二组参数最大迭代次数（None时从配置读取）
    lambda_init_group1: 第一组参数初始阻尼因子（None时从配置读取）
    lambda_init_group2: 第二组参数初始阻尼因子（None时从配置读取）
    max_theta_delta_rad_for_sub_opt: theta参数最大变化量（None时从配置读取）
    """
    # 从配置文件读取默认参数
    alt_config = get_alternate_optimization_config()
    damping_config = get_damping_config()
    output_config = get_output_config()
    
    # 使用传入参数或配置文件中的默认值
    if max_alt_iterations is None:
        max_alt_iterations = alt_config.get('max_alt_iterations', 4)
    if convergence_tol is None:
        convergence_tol = alt_config.get('convergence_tol', 1e-4)
    if max_sub_iterations_group1 is None:
        max_sub_iterations_group1 = alt_config.get('max_sub_iterations_group1', 10)
    if max_sub_iterations_group2 is None:
        max_sub_iterations_group2 = alt_config.get('max_sub_iterations_group2', 10)
    if lambda_init_group1 is None:
        lambda_init_group1 = damping_config.get('lambda_init_group1', 2.0)
    if lambda_init_group2 is None:
        lambda_init_group2 = damping_config.get('lambda_init_group2', 0.001)
    if max_theta_delta_rad_for_sub_opt is None:
        max_theta_delta_rad_for_sub_opt = get_max_theta_change_radians()
    
    print("\n" + "="*60)
    print(" "*20 + "开始交替优化")
    print("="*60)
    
    # 从配置文件读取CSV文件路径
    csv_file = None
    if output_config.get('save_delta_values', True):
        csv_file = output_config.get('delta_csv_file', 'results/delta_values.csv')
    
    # 创建参数名称列表
    param_names = []
    for i in range(6):
        for param in ["alpha", "a", "d", "theta_offset"]:
            param_names.append(f"Joint{i+1}_{param}")
    for param in ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]:
        param_names.append(f"TCP_{param}")
    for param in ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]:
        param_names.append(f"Laser_{param}")
    
    # 初始化CSV文件
    if csv_file:
        try:
            # 创建目录（如果不存在）
            csv_dir = os.path.dirname(csv_file)
            if csv_dir and not os.path.exists(csv_dir):
                os.makedirs(csv_dir)
                
            with open(csv_file, 'w', newline='') as f:
                writer = csv.writer(f)
                # 只包含参数名称，不包含Iteration、Lambda和Error
                header = param_names
                writer.writerow(header)
            print(f"Delta值将保存到: {csv_file}")
        except Exception as e:
            print(f"创建CSV文件时出错: {e}")
            csv_file = None
    
    #* 读取关节角度和激光数据
    joint_angles = load_joint_angles()
    laser_matrices = get_laser_tool_matrix()
    
    #! 获取参数分组索引
    opt_indices_group1, opt_indices_group2 = get_parameter_groups(exclude_fixed=True)
    
    #! 初始化均方根误差
    params = np.array(initial_params)
    current_avg_error_val = compute_total_error_avg(params, joint_angles, laser_matrices).item() # 已修改为计算平均误差, 使用 .item() 获取数值
    avg_initial_error_alternate = current_avg_error_val 
    print(f"初始均方根误差: {avg_initial_error_alternate:.6f}")
    
    #* 打印参数组信息
    print(f"第一组参数索引 (共{len(opt_indices_group1)}个): {opt_indices_group1}")
    print(f"第二组参数索引 (共{len(opt_indices_group2)}个): {opt_indices_group2}")
    
    # 记录误差历史
    error_history_avg = [avg_initial_error_alternate] # 记录平均误差, 使用新的初始误差变量名
    
    #! 交替优化主循环
    for alt_iteration in range(max_alt_iterations):
        print(f"\n===== 交替优化循环 {alt_iteration + 1}/{max_alt_iterations} =====")
        
        #! 第一步：优化DH参数 + TCP + 激光XYZ，固定激光四元数
        print("\n----- 第一步：优化DH参数 + TCP + 激光XYZ -----")
        params_step1 = optimize_dh_parameters(
            params, 
            max_iterations=max_sub_iterations_group1, # 修改：使用第一组的迭代次数
            lambda_init=lambda_init_group1, 
            opt_indices=opt_indices_group1,
            max_theta_delta_rad=max_theta_delta_rad_for_sub_opt,
            csv_file=csv_file,
            alt_iteration=alt_iteration+1,
            opt_step=1
        )
        
        #* 计算第一步优化后的误差    
        avg_error_step1 = compute_total_error_avg(params_step1, joint_angles, laser_matrices).item() 
        print(f"第一步后均方根误差: {avg_error_step1:.6f}")
        
        #! 第二步：优化激光四元数，固定DH参数+TCP+激光XYZ
        print("\n----- 第二步：优化激光四元数 -----")
        params_step2 = optimize_dh_parameters(
            params_step1, 
            max_iterations=max_sub_iterations_group2, # 修改：使用第二组的迭代次数
            lambda_init=lambda_init_group2, 
            opt_indices=opt_indices_group2,
            max_theta_delta_rad=max_theta_delta_rad_for_sub_opt,
            csv_file=csv_file,
            alt_iteration=alt_iteration+1,
            opt_step=2
        )
        
        #* 计算第二步优化后的误差
        avg_error_step2 = compute_total_error_avg(params_step2, joint_angles, laser_matrices).item() 
        print(f"第二步后均方根误差: {avg_error_step2:.6f}")
        
        #* 更新误差
        params = params_step2
        error_history_avg.append(avg_error_step2) 
        
        #* 计算误差改进量
        error_improvement = error_history_avg[-2] - error_history_avg[-1] 
        relative_improvement = error_improvement / error_history_avg[-2] if error_history_avg[-2] > 1e-9 else 0
        
        print(f"\n本次循环误差改进: {error_improvement:.6f}, 相对改进: {relative_improvement*100:.4f}%")
        
        if error_improvement < convergence_tol:
            print(f"\n误差改进 {error_improvement:.6f} 小于阈值 {convergence_tol}，交替优化收敛")
            break
            
    #* 计算总体优化效果
    final_avg_error_alternate = error_history_avg[-1] 
    total_improvement = (avg_initial_error_alternate - final_avg_error_alternate) / avg_initial_error_alternate * 100 if avg_initial_error_alternate > 1e-9 else 0
    
    print("\n" + "="*60)
    print(f"交替优化完成，共进行了 {alt_iteration + 1} 次循环")
    print(f"初始均方根误差: {avg_initial_error_alternate:.6f}") 
    print(f"最终均方根误差: {final_avg_error_alternate:.6f}")
    print(f"总体改进率: {total_improvement:.2f}%")
    print("="*60)
    
    return params

def evaluate_optimization(initial_params, optimized_params):
    """评估优化效果，报告与优化器目标一致的均方根误差"""
    # 读取数据
    joint_angles = load_joint_angles()
    laser_matrices = get_laser_tool_matrix()
    n_samples = len(joint_angles)

    if n_samples == 0:
        print("评估警告: 样本数量为0，无法进行评估。")
        return
    
    print("\n" + "="*60) # 调整分隔线长度以适应新的表头 
    print(" "*15 + "优化效果评估 (所有分量的均方根误差)") # 修改标题
    print("="*60)
    # 打印表头，明确指出总体平均误差是均方根误差
    print(f"{'姿态(帧)':^12}|{'初始单帧范数':^18}|{'优化后单帧范数':^20}|{'单帧改进率':^15}")
    print("-"*68) # 调整分隔线长度
    
    # 计算初始和优化后的总体均方根误差 (与compute_total_error_avg一致)
    initial_total_rmse = compute_total_error_avg(initial_params, joint_angles, laser_matrices).item()
    optimized_total_rmse = compute_total_error_avg(optimized_params, joint_angles, laser_matrices).item()

    # 逐帧显示误差范数及其改进，用于详细分析
    for i in range(n_samples):
        initial_error_vec = compute_error_vector(initial_params, joint_angles[i], laser_matrices[i])
        optimized_error_vec = compute_error_vector(optimized_params, joint_angles[i], laser_matrices[i])
        
        initial_frame_norm = torch.linalg.norm(initial_error_vec).item()
        optimized_frame_norm = torch.linalg.norm(optimized_error_vec).item()
        
        frame_improvement = (1 - optimized_frame_norm / initial_frame_norm) * 100 if initial_frame_norm > 1e-9 else 0
        
        print(f"{i+1:^12}|{initial_frame_norm:^18.6f}|{optimized_frame_norm:^20.6f}|{frame_improvement:^14.2f}%")
    
    # 计算总体改进率 (基于均方根误差)
    avg_improvement_rmse = (1 - optimized_total_rmse / initial_total_rmse) * 100 if initial_total_rmse > 1e-9 else 0
    
    print("-"*68) # 调整分隔线长度
    print(f"{'总体平均RMSE':^12}|{initial_total_rmse:^18.6f}|{optimized_total_rmse:^20.6f}|{avg_improvement_rmse:^14.2f}%")
    print("="*60)


if __name__ == '__main__':
    initial_params = get_initial_params()

    # 获取可优化参数索引
    opt_indices = get_optimizable_indices()
    fixed_indices = get_fixed_indices()  # 直接调用函数获取
    print(f"固定参数索引 ({len(fixed_indices)}): {fixed_indices}")
    print(f"可优化参数索引 ({len(opt_indices)}): {opt_indices}")
    
    # 从配置文件读取theta参数最大变化量
    max_theta_change_radians = get_max_theta_change_radians()
    max_theta_change_degrees = np.rad2deg(max_theta_change_radians)
    print(f"Theta参数单步最大变化量: {max_theta_change_degrees:.1f}度 ({max_theta_change_radians:.6f}弧度)")
    
    # 使用交替优化方法 - 现在所有参数都从配置文件读取
    optimized_params = alternate_optimize_parameters(initial_params)

    # 保存优化结果 - 路径从配置文件读取
    save_optimization_results(optimized_params, initial_params) 

    # 评估优化效果 
    evaluate_optimization(initial_params, optimized_params)
    
    # 输出优化前后的参数对比
    print("\n" + "="*70)
    print(" "*25 + "DH参数对比")
    print("="*70)
    print(f"{'关节':^6}|{'参数':^12}|{'初始值':^15}|{'优化值':^15}|{'差异':^15}|{'状态':^10}")
    print("-"*70)
    
    param_names = ["alpha", "a", "d", "theta_offset"]
    
    # 将参数重构为6×4矩阵，方便查看 (DH部分)
    init_dh_matrix = initial_params[0:24].reshape(6, 4)
    opt_dh_matrix = optimized_params[0:24].reshape(6, 4)
    
    for i in range(6):  
        for j in range(4):  
            param_idx = i * 4 + j  
            param_diff = opt_dh_matrix[i, j] - init_dh_matrix[i, j]
            status = "已优化" if param_idx in opt_indices else "已固定"
            print(f"{i+1:^6}|{param_names[j]:^12}|{init_dh_matrix[i, j]:^15.4f}|{opt_dh_matrix[i, j]:^15.4f}|{param_diff:^15.4f}|{status:^10}")
        if i < 5:  
            print("-"*70)
    
    # 添加TCP参数对比
    print("="*70)
    print(" "*25 + "TCP 参数对比")
    print("="*70)
    tcp_param_names = ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]
    init_tcp_params = initial_params[24:31]
    opt_tcp_params = optimized_params[24:31]
    for k in range(7):
        tcp_idx = 24 + k
        tcp_diff = opt_tcp_params[k] - init_tcp_params[k]
        status = "已优化" if tcp_idx in opt_indices else "已固定"
        print(f"{'-':^6}|{tcp_param_names[k]:^12}|{init_tcp_params[k]:^15.4f}|{opt_tcp_params[k]:^15.4f}|{tcp_diff:^15.4f}|{status:^10}")

    # 添加激光跟踪仪-基座变换参数对比
    print("="*70)
    print(" "*25 + "激光跟踪仪-基座变换参数对比")
    print("="*70)
    t_laser_base_param_names = ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]
    init_t_laser_base_params = initial_params[31:38]
    opt_t_laser_base_params = optimized_params[31:38]
    for k in range(7):
        t_laser_base_idx = 31 + k
        t_laser_base_diff = opt_t_laser_base_params[k] - init_t_laser_base_params[k]
        status = "已优化" if t_laser_base_idx in opt_indices else "已固定"
        print(f"{'-':^6}|{t_laser_base_param_names[k]:^12}|{init_t_laser_base_params[k]:^15.4f}|{opt_t_laser_base_params[k]:^15.4f}|{t_laser_base_diff:^15.4f}|{status:^10}")

    print("="*70)