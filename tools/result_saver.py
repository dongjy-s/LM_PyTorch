"""
机器人标定优化结果保存模块

包含以下功能：
- 保存参数更新量到CSV文件
- 保存优化结果（DH参数、TCP参数、激光跟踪仪参数）
- 生成Excel汇总报告
- 支持详细对比分析
"""

import os
import csv
import numpy as np
from .data_loader import get_output_config

def save_delta_to_csv(delta, iteration, opt_indices, csv_file, lambda_val=None, error_val=None, alt_iteration=None, opt_step=None):
    """
    将delta值保存到CSV文件
    
    参数:
    delta: 参数更新量
    iteration: 当前迭代次数
    opt_indices: 可优化参数的索引
    csv_file: CSV文件路径
    lambda_val: 当前阻尼因子值（可选）
    error_val: 当前误差值（可选）
    alt_iteration: 交替优化循环次数（可选）
    opt_step: 优化步骤（1或2，可选）
    """
    try:
        # 创建目录（如果不存在）
        csv_dir = os.path.dirname(csv_file)
        if csv_dir and not os.path.exists(csv_dir):
            os.makedirs(csv_dir)
            
        # 创建完整的delta数组（38个参数）
        full_delta = np.zeros(38)
        for i, idx in enumerate(opt_indices):
            full_delta[idx] = delta[i]
        
        # 写入CSV文件
        with open(csv_file, 'a', newline='') as f:
            writer = csv.writer(f)
            row = []
            # 只添加delta值，不添加迭代信息、lambda和误差值
            row.extend(full_delta)
            writer.writerow(row)
    except Exception as e:
        print(f"保存delta值到CSV文件时出错: {e}")

def save_optimization_results(params, initial_params=None, filepath_prefix=None):
    """
    保存优化结果，路径可以从配置文件读取
    
    参数:
    params: 优化后的参数
    initial_params: 初始参数（用于对比分析）
    filepath_prefix: 文件路径前缀（None时从配置读取）
    """
    # 从配置文件读取默认路径
    if filepath_prefix is None:
        output_config = get_output_config()
        filepath_prefix = output_config.get('results_prefix', 'results/optimized')
    
    dirpath = os.path.dirname(filepath_prefix)
    if dirpath and not os.path.exists(dirpath):
        os.makedirs(dirpath)
    
    # 分离不同类型的参数
    dh_params = params[0:24]
    tcp_params = params[24:31]
    t_laser_base_params = params[31:38]
    
    # 保存DH参数
    dh_filepath = f"{filepath_prefix}_dh_parameters.csv"
    dh_matrix = np.array(dh_params).reshape(6, 4)
    header_dh = "alpha,a,d,theta_offset"
    row_labels_dh = [f"Joint_{i+1}" for i in range(6)]
    with open(dh_filepath, 'w') as f:
        f.write(f",{header_dh}\n")  
        for i, row in enumerate(dh_matrix):
            f.write(f"{row_labels_dh[i]},{row[0]:.6f},{row[1]:.6f},{row[2]:.6f},{row[3]:.6f}\n")
    print(f"优化后的DH参数已保存到: {dh_filepath}")
    
    # 保存TCP参数
    tcp_filepath = f"{filepath_prefix}_tcp_parameters.csv"
    header_tcp = "parameter,value"
    tcp_param_names = ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]
    with open(tcp_filepath, 'w') as f:
        f.write(f"{header_tcp}\n")
        for name, value in zip(tcp_param_names, tcp_params):
            f.write(f"{name},{value:.6f}\n")
    print(f"优化后的TCP参数已保存到: {tcp_filepath}")
    
    # 保存激光跟踪仪-基座变换参数
    t_laser_base_filepath = f"{filepath_prefix}_t_laser_base_parameters.csv"
    header_t_laser_base = "parameter,value"
    t_laser_base_param_names = ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]
    with open(t_laser_base_filepath, 'w') as f:
        f.write(f"{header_t_laser_base}\n")
        for name, value in zip(t_laser_base_param_names, t_laser_base_params):
            f.write(f"{name},{value:.6f}\n")
    print(f"优化后的激光跟踪仪-基座变换参数已保存到: {t_laser_base_filepath}")
    
    # 生成详细对比分析数据
    detailed_comparison_df = None
    if initial_params is not None:
        try:
            from .result_analyzer import generate_detailed_comparison
            print("\n🔍 生成优化前后详细对比分析...")
            detailed_comparison_df = generate_detailed_comparison(initial_params, params)
        except ImportError:
            print("警告: 无法导入结果分析模块，跳过详细对比分析")
    
    # 生成Excel汇总文件
    _generate_excel_summary(
        dh_filepath, tcp_filepath, t_laser_base_filepath, 
        dirpath, detailed_comparison_df
    )

def _generate_excel_summary(dh_filepath, tcp_filepath, t_laser_base_filepath, 
                          results_dir, detailed_comparison_df):
    """
    生成Excel汇总文件的内部函数
    """
    try:
        import pandas as pd
        from pathlib import Path
        
        # 创建Excel汇总文件
        results_dir_path = Path(results_dir)
        excel_filepath = results_dir_path / "优化结果汇总.xlsx"
        
        # 读取三个CSV文件的数据
        dh_df = pd.read_csv(dh_filepath)
        tcp_df = pd.read_csv(tcp_filepath)
        base_df = pd.read_csv(t_laser_base_filepath)
        
        # 创建一个综合的数据框，将三个结果合并到一个工作表
        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
            # 创建工作簿
            workbook = writer.book
            worksheet = workbook.create_sheet(title="优化结果汇总")
            
            # 设置样式
            _setup_excel_styles(worksheet, workbook)
            
            # 写入各个部分的数据
            row = _write_dh_parameters_section(worksheet, dh_df)
            row = _write_tcp_section(worksheet, tcp_df, row)
            row = _write_base_section(worksheet, base_df, row)
            _write_explanation_section(worksheet, row)
            
            # 添加详细对比分析工作表
            if detailed_comparison_df is not None:
                _create_comparison_worksheet(workbook, detailed_comparison_df)
            
            # 删除默认的Sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        print(f"✅ 优化结果汇总Excel文件已生成: {excel_filepath}")
        _print_excel_summary()
        
    except ImportError:
        print("⚠️  未安装pandas库，无法生成Excel汇总文件")
        print("   请运行: pip install pandas openpyxl")
    except Exception as e:
        print(f"⚠️  生成Excel汇总文件时出错: {e}")

def _setup_excel_styles(worksheet, workbook):
    """设置Excel样式"""
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # 调整列宽
    worksheet.column_dimensions['A'].width = 18
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 15

def _write_dh_parameters_section(worksheet, dh_df):
    """写入DH参数部分"""
    from openpyxl.styles import Font, Alignment, PatternFill
    
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    # 写入DH参数部分
    row = 1
    worksheet.cell(row=row, column=1, value="🔧 机器人DH参数标定结果").font = title_font
    worksheet.cell(row=row, column=1).alignment = center_alignment
    worksheet.merge_cells(f'A{row}:E{row}')
    row += 2
    
    # DH参数表头
    dh_headers = ['关节编号', 'α (度)', 'a (mm)', 'd (mm)', 'θ偏移 (度)']
    for col, header in enumerate(dh_headers, 1):
        cell = worksheet.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    row += 1
    
    # DH参数数据
    for i, (_, dh_row) in enumerate(dh_df.iterrows()):
        worksheet.cell(row=row, column=1, value=dh_row.iloc[0])  # 关节名称
        for col in range(1, 5):
            worksheet.cell(row=row, column=col+1, value=f"{dh_row.iloc[col]:.6f}")
        row += 1
    
    return row + 2  # 空行分隔

def _write_tcp_section(worksheet, tcp_df, start_row):
    """写入TCP位姿部分"""
    from openpyxl.styles import Font, Alignment, PatternFill
    
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    row = start_row
    # 写入TCP位姿部分
    worksheet.cell(row=row, column=1, value="🎯 工具中心点(TCP)位姿").font = title_font
    worksheet.cell(row=row, column=1).alignment = center_alignment
    worksheet.merge_cells(f'A{row}:B{row}')
    row += 2
    
    # TCP位姿表头和数据
    tcp_headers = ['位姿参数', '优化值']
    for col, header in enumerate(tcp_headers, 1):
        cell = worksheet.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    row += 1
    
    # TCP参数名称美化映射
    tcp_param_names = {
        'tx': 'X轴位移 (mm)',
        'ty': 'Y轴位移 (mm)', 
        'tz': 'Z轴位移 (mm)',
        'qx': '四元数 qx',
        'qy': '四元数 qy',
        'qz': '四元数 qz',
        'qw': '四元数 qw'
    }
    
    for _, tcp_row in tcp_df.iterrows():
        param_name = tcp_param_names.get(tcp_row['parameter'], tcp_row['parameter'])
        worksheet.cell(row=row, column=1, value=param_name)
        worksheet.cell(row=row, column=2, value=f"{tcp_row['value']:.6f}")
        row += 1
    
    return row

def _write_base_section(worksheet, base_df, start_row):
    """写入基座位姿部分"""
    from openpyxl.styles import Font, Alignment, PatternFill
    
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    # 写入基座位姿部分（与TCP位姿并排显示）
    row_start_base = start_row - len(base_df) - 1  # 回到TCP标题行
    worksheet.cell(row=row_start_base, column=4, value="📍 激光跟踪仪基座位姿").font = title_font
    worksheet.cell(row=row_start_base, column=4).alignment = center_alignment
    worksheet.merge_cells(f'D{row_start_base}:E{row_start_base}')
    row_base = row_start_base + 2
    
    # 基座位姿表头
    base_headers = ['位姿参数', '优化值']
    for col, header in enumerate(base_headers, 4):
        cell = worksheet.cell(row=row_base, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    row_base += 1
    
    # 基座参数名称美化映射
    base_param_names = {
        'tx': 'X轴位移 (mm)',
        'ty': 'Y轴位移 (mm)',
        'tz': 'Z轴位移 (mm)', 
        'qx': '四元数 qx',
        'qy': '四元数 qy',
        'qz': '四元数 qz',
        'qw': '四元数 qw'
    }
    
    # 基座位姿数据
    for _, base_row in base_df.iterrows():
        param_name = base_param_names.get(base_row['parameter'], base_row['parameter'])
        worksheet.cell(row=row_base, column=4, value=param_name)
        worksheet.cell(row=row_base, column=5, value=f"{base_row['value']:.6f}")
        row_base += 1
    
    return max(start_row, row_base)

def _write_explanation_section(worksheet, start_row):
    """写入说明部分"""
    from openpyxl.styles import Font, Alignment
    
    title_font = Font(bold=True, size=14)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 在主工作表底部添加说明
    last_row = start_row + 2
    worksheet.cell(row=last_row, column=1, value="📋 说明").font = title_font
    worksheet.cell(row=last_row, column=1).alignment = center_alignment
    
    explanation_lines = [
        "• DH参数：机器人正向运动学标定参数",
        "• TCP位姿：工具中心点相对法兰坐标系的变换",
        "• 基座位姿：激光跟踪仪与机器人基座间的变换关系",
        "• 详细误差分析请查看'优化前后详细对比'工作表"
    ]
    
    for i, line in enumerate(explanation_lines):
        worksheet.cell(row=last_row + 1 + i, column=1, value=line).font = Font(size=10)
        worksheet.merge_cells(f'A{last_row + 1 + i}:E{last_row + 1 + i}')
    
    # 添加边框样式
    _apply_borders_to_worksheet(worksheet)

def _apply_borders_to_worksheet(worksheet):
    """为工作表添加边框"""
    from openpyxl.styles import Border, Side
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 为所有有数据的单元格添加边框
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

def _create_comparison_worksheet(workbook, detailed_comparison_df):
    """创建对比分析工作表"""
    from openpyxl.styles import Font, Alignment, PatternFill
    
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    
    # 创建对比分析工作表
    comparison_sheet = workbook.create_sheet(title="优化前后详细对比")
    
    # 写入标题
    comparison_sheet.cell(row=1, column=1, value="📊 优化前后逐组数据对比分析").font = title_font
    comparison_sheet.cell(row=1, column=1).alignment = center_alignment
    comparison_sheet.merge_cells('A1:V1')  # 扩展合并范围以适应新列
    
    # 添加计算说明
    explanation_text = ("💡 误差计算方式:\n"
                       "   位置误差 = √(X²+Y²+Z²) (mm)\n"
                       "   姿态误差 = √(Rx²+Ry²+Rz²) (度)\n" 
                       "   总误差(L2范数) = √(X²+Y²+Z²+(Rx×0.01)²+(Ry×0.01)²+(Rz×0.01)²)\n"
                       "   权重设置: 位置权重=1.0, 姿态权重=0.01")
    comparison_sheet.cell(row=2, column=1, value=explanation_text).font = Font(size=10, italic=True)
    comparison_sheet.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    comparison_sheet.merge_cells('A2:V2')  # 扩展合并范围以适应新列
    comparison_sheet.row_dimensions[2].height = 40
    
    # 写入表头
    headers = detailed_comparison_df.columns.tolist()
    for col, header in enumerate(headers, 1):
        cell = comparison_sheet.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # 写入数据
    for row_idx, (_, row_data) in enumerate(detailed_comparison_df.iterrows(), 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = comparison_sheet.cell(row=row_idx, column=col_idx, value=value)
            # 为平均值行添加特殊样式
            if '平均值' in str(value):
                cell.font = header_font
                cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        if col == 1:  # 数据组列
            comparison_sheet.column_dimensions[chr(64 + col)].width = 12
        elif '改进率' in headers[col-1]:  # 改进率列
            comparison_sheet.column_dimensions[chr(64 + col)].width = 14
        elif '位置误差' in headers[col-1] or '姿态误差' in headers[col-1]:  # 位置/姿态误差列
            comparison_sheet.column_dimensions[chr(64 + col)].width = 18
        else:  # 其他误差列
            comparison_sheet.column_dimensions[chr(64 + col)].width = 16
    
    _apply_borders_to_worksheet(comparison_sheet)

def _print_excel_summary():
    """打印Excel汇总信息"""
    print("📊 Excel文件包含完整的优化分析:")
    print("   🔧 机器人DH参数标定结果 - 6个关节的完整DH参数")
    print("   🎯 工具中心点(TCP)位姿 - 位移+四元数表示") 
    print("   📍 激光跟踪仪基座位姿 - 基座坐标系变换参数")
    print("   📊 优化前后详细对比 - 逐组数据误差分析和改进率")
    print("   ✨ 专业格式化：图标标识、单位标注、边框美化")
    print("   📋 多工作表展示，全面分析优化效果")

def initialize_delta_csv(csv_file):
    """
    初始化delta值CSV文件
    
    参数:
    csv_file: CSV文件路径
    """
    # 创建参数名称列表
    param_names = []
    for i in range(6):
        for param in ["alpha", "a", "d", "theta_offset"]:
            param_names.append(f"Joint{i+1}_{param}")
    for param in ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]:
        param_names.append(f"TCP_{param}")
    for param in ["tx", "ty", "tz", "qx", "qy", "qz", "qw"]:
        param_names.append(f"Laser_{param}")
    
    try:
        # 创建目录（如果不存在）
        csv_dir = os.path.dirname(csv_file)
        if csv_dir and not os.path.exists(csv_dir):
            os.makedirs(csv_dir)
            
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            # 只包含参数名称
            header = param_names
            writer.writerow(header)
        print(f"Delta值CSV文件已初始化: {csv_file}")
        return True
    except Exception as e:
        print(f"初始化CSV文件时出错: {e}")
        return False 